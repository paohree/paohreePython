import numpy as np
import matplotlib.pyplot as plt
import openpyxl

file = openpyxl.load_workbook('data_set_train.xlsx')
ws = file.active

data = []
col_name = []
for row in ws.iter_rows(max_row=1):
    for cell in row:
        col_name.append(cell.value)
for row in ws.iter_rows(min_row=2):
    one_line = []
    for cell in row:
        one_line.append(cell.value)
    data.append(one_line)

# print(type(data))
arr = np.array(data)
# print(type(arr))
# print(arr.shape)
data_set = arr[:, :]
print("Exercise 1")


def sum(n):
    if n == 0:
        return 0
    return n + sum(n - 1)


result = sum(10)
print(result)


def is_odd(num):
    if num % 2 == 0:
        return "Even"
    else:
        return "Odd"


result = is_odd(5)
print(result)
print("")

print("exercise 2")
data_set2006 = data_set[data_set[:, 1] < '2007Q1', :]
print(data_set2006)
print("")

print("Exercise 3")
station = arr[:, 19].astype(float)
price = arr[:, 2].astype(float)
print(np.corrcoef(price, station))
print("")

print("exercise 4")


def price(version, year):
    if version == 1:
        if year == 2006:
            index = str(year + 1) + 'Q' + '1'
            return np.mean(arr[data_set[:, 1] < index, 2].astype(float))
        if year == 2017:
            index = str(year - 1) + 'Q' + '4'
            return np.mean(arr[data_set[:, 1] > index, 2].astype(float))
    else:
        first_index = str(year - 1) + 'Q' + '4'
        second_index = str(year + 1) + 'Q' + '1'
        return np.mean(arr[(data_set[:, 1] > first_index) & (data_set[:, 1] < second_index), 2].astype(float))


for year in range(2006, 2018):
    if year == 2006 or year == 2017:
        mean_price = price(1, year)
        formatted_price = "{:.2f}".format(mean_price)
        print(str(year) + " avg : " + formatted_price)
    else:
        mean_price = price(2, year)
        formatted_price = "{:.2f}".format(mean_price)
        print(str(year) + " avg : " + formatted_price)

print("")

print("Exercise 5")


def priceVer2(quarter, year):
    if quarter == 1:
        if year == 2006:
            index = str(year) + 'Q' + str(quarter + 1)  # 2006Q2가 되겠지
            return np.mean(arr[data_set[:, 1] < index, 2].astype(float))
        else:
            first_index = str(year - 1) + 'Q' + str(quarter + 3)
            second_index = str(year) + 'Q' + str(quarter + 1)
            return np.mean(arr[(data_set[:, 1] > first_index) & (data_set[:, 1] < second_index), 2].astype(float))
    elif quarter == 2 or quarter == 3:
        first_index = str(year) + 'Q' + str(quarter - 1)
        second_index = str(year) + 'Q' + str(quarter + 1)
        return np.mean(arr[(data_set[:, 1] > first_index) & (data_set[:, 1] < second_index), 2].astype(float))
    else:
        if year == 2017:
            index = str(year) + 'Q' + str(quarter - 1)
            return np.mean(arr[data_set[:, 1] > index, 2].astype(float))
        else:
            first_index = str(year) + 'Q' + str(quarter - 1)
            second_index = str(year + 1) + 'Q' + str(quarter - 3)
            return np.mean(arr[(data_set[:, 1] > first_index) & (data_set[:, 1] < second_index), 2].astype(float))


for year in range(2006, 2018):
    for quarter in range(1, 5):
        if quarter==4 and year==2017:
            break
        mean_price = priceVer2(int(quarter), int(year))
        formatted_price = "{:.2f}".format(mean_price)
        if quarter == 4:
            print(str(year) + "Q" + str(quarter) + " avg : " + formatted_price)
        else:
            print(str(year) + "Q" + str(quarter) + " avg : " + formatted_price, end="  ")
print("")

print("Exercise 6")

import numpy as np
import matplotlib.pyplot as plt

quarters = arr[:, 1]
prices = arr[:, 2].astype(float)

unique_quarters, counts = np.unique(quarters, return_counts=True)

average_prices = []
for quarter in unique_quarters:
    mask = quarters == quarter
    average_price = np.mean(prices[mask])
    average_prices.append(average_price)

fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 10))

ax1.bar(unique_quarters, counts)
ax1.set_title("Count of Trading in Quarters")
ax1.set_xlabel("Quarters")
ax1.set_ylabel("Counts")
plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45)

ax2.plot(unique_quarters, average_prices, marker='o', linestyle='-')
ax2.set_title("Quarterly Average Price")
ax2.set_xlabel("Quarters")
ax2.set_ylabel("Average Price")
plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45)

plt.tight_layout()
plt.show()

print("")